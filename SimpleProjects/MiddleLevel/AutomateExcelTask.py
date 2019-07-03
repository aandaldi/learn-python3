import openpyxl

#open workbook
workbook = openpyxl.load_workbook("Employees.xlsx")
sheet = workbook['EmployeeData']
print(workbook.sheetnames, sheet)                              # to see sheetnames
print(workbook.active)

#create sheet
# # workbook.create_sheet("TestCreate")
# workbook.save("Employess.xlsx")
# sheet = workbook['TestCreate']
# workbook.remove(sheet)
# del workbook['TestCreate']

print(sheet.dimensions)
print(sheet.cell(row = 6, column = 2).value)

## show all value from sheet
# for i in sheet.values:
#     # print(i)
#     print(i[4])

#using style on woorkbook
from openpyxl.styles import  *





